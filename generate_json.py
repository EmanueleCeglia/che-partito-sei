# OBSOLETO — generatore della v1 del quiz. NON eseguire.
#
# Produce lo schema vecchio (categories -> themes -> statements, una frase per
# partito), mentre data.json usa ormai quello della v2: categories -> questions,
# ognuna con "scores" per gli 11 partiti. Rilanciarlo sovrascriverebbe il quiz
# attuale con 297 domande e i nomi partiti non normalizzati.
#
# data.json e' versionato: la fonte di verita' e' il repo, non questo script.
# Resta qui solo come traccia di come i dati furono estratti dall'Excel
# (Politics_Leo_Ema_Dani.xlsx, gitignorato).

import openpyxl
import json

wb = openpyxl.load_workbook(r'd:\Desktop\Project Politics\Politics_Leo_Ema_Dani.xlsx', data_only=True)
ws = wb['Score']

# Macro-temi come specificati dall'utente (nomi esatti dal prompt)
macro_themes_map = {
    4: "Economia, Fisco e Lavoro",    # D
    5: "Economia, Fisco e Lavoro",    # E
    6: "Economia, Fisco e Lavoro",    # F
    7: "Economia, Fisco e Lavoro",    # G
    8: "Economia, Fisco e Lavoro",    # H
    9: "Welfare, Salute e Istruzione",    # I
    10: "Welfare, Salute e Istruzione",   # J
    11: "Welfare, Salute e Istruzione",   # K
    12: "Welfare, Salute e Istruzione",   # L
    13: "Diritti Civili, Etica e Società",  # M
    14: "Diritti Civili, Etica e Società",  # N
    15: "Diritti Civili, Etica e Società",  # O
    16: "Diritti Civili, Etica e Società",  # P
    17: "Esteri",    # Q
    18: "Esteri",    # R
    19: "Esteri",    # S
    20: "Esteri",    # T
    21: "Trans.Ecologica ed Energia",  # U
    22: "Trans.Ecologica ed Energia",  # V
    23: "Trans.Ecologica ed Energia",  # W
    24: "Sicurezza",   # X
    25: "Sicurezza",   # Y
    26: "Sicurezza",   # Z
    27: "Istituzioni, Democrazia e PA",  # AA
    28: "Istituzioni, Democrazia e PA",  # AB
    29: "Istituzioni, Democrazia e PA",  # AC
    30: "Istituzioni, Democrazia e PA",  # AD
}

# Leggi i partiti (riga 3-13, colonna C)
parties = []
for row in range(3, 14):
    party_name = ws.cell(row=row, column=3).value
    if party_name:
        parties.append(party_name.strip())

# Leggi i temi (riga 2, colonne D-AD)
themes = {}
for col in range(4, 31):
    theme_name = ws.cell(row=2, column=col).value
    if theme_name:
        themes[col] = theme_name.strip()

# Costruisci la struttura JSON
categories_dict = {}

for col in range(4, 31):
    macro_theme = macro_themes_map.get(col)
    theme_name = themes.get(col)
    
    if not macro_theme or not theme_name:
        continue
    
    if macro_theme not in categories_dict:
        categories_dict[macro_theme] = {
            "name": macro_theme,
            "themes": []
        }
    
    theme_data = {
        "name": theme_name,
        "statements": []
    }
    
    for row_idx, party in enumerate(parties):
        cell_value = ws.cell(row=row_idx + 3, column=col).value
        if cell_value:
            statement_text = str(cell_value).strip()
            theme_data["statements"].append({
                "party": party,
                "text": statement_text
            })
    
    categories_dict[macro_theme]["themes"].append(theme_data)

# Ordine dei macro-temi come nel prompt
macro_order = [
    "Economia, Fisco e Lavoro",
    "Welfare, Salute e Istruzione",
    "Diritti Civili, Etica e Società",
    "Esteri",
    "Trans.Ecologica ed Energia",
    "Sicurezza",
    "Istituzioni, Democrazia e PA"
]

categories = [categories_dict[m] for m in macro_order if m in categories_dict]

data = {
    "parties": parties,
    "categories": categories
}

# Conta totale domande
total = sum(
    len(t["statements"]) 
    for c in categories 
    for t in c["themes"]
)
print(f"Partiti: {len(parties)}")
print(f"Categorie: {len(categories)}")
print(f"Temi totali: {sum(len(c['themes']) for c in categories)}")
print(f"Domande totali: {total}")

# Salva JSON
with open(r'd:\Desktop\Project Politics\data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print("\nJSON salvato in data.json")
